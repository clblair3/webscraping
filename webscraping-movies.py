
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2024/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)
##
##
##
##

wb = xl.Workbook()
ws = wb.active

ws.title = 'Box Office Report'

#headers
ws['A1'] = 'No.'
ws['B1'] = 'Movie Title'
ws['C1'] = 'Gross'
ws['D1'] = 'Theaters'
ws['E1'] = 'Average Gross/Theater'


#give us a long list of tr tags
movie_rows = soup.findAll('tr')


#in the second tr tag, there are abunch of td tags
for x in range (1, 6):
    td = movie_rows[x].findAll('td')

    rank = (td[0].text)
    name = (td[1].text)
    gross = int((td[5].text).replace("$", "").replace(",","")) #used 5 and 6 because there were blank spaces in the code
    theaters = int((td[6].text).replace(",",""))

    avg = gross/theaters

    #writing this to the excel file
    ws['A' + str(x+1)] = rank
    ws['B' + str(x+1)] = name
    ws['C' + str(x+1)] = gross
    ws['D' + str(x+1)] = theaters
    ws['E' + str(x+1)] = avg

header_font = Font(size=16, bold=True)

for cell in ws[1:1]:
    cell.font = header_font

for cell in ws['D:D']:
    cell.number_format = '#,##0.00'

for cell in ws['B:B']:
    cell.number_format = u'"$ "#,##0.00'

for cell in ws['E:E']:
    cell.number_format = u'"$ "#,##0.00'

#change column size
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 35




#save workbook
wb.save('BoxOfficeReport.xlsx')